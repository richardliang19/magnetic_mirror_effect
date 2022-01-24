"""
 Vpython 磁瓶效應
 資料來源: 
 http://keejko.blogspot.com/2018/05/blog-post_14.html

"""
# 引入套件
from vpython import *
import numpy as np
import pyscreenshot as ImageGrab
import openpyxl 

"""
 1. 參數設定, 設定變數及初始值
"""
size = 0.4            # 螺線管截面的半徑
point_size = 0.2*size # 螺線管導線分割後標示位置用的小球半徑, 若要使小球較明顯設為1倍, 若要隱藏設為0.2倍
seg_size = 1.0*size   # 螺線管導線分割後每一小段導線的半徑, 若要使導線較明顯設為1倍, 若要隱藏設為0.2倍
r = 10                # 螺線管的半徑
n = 800               # 螺線管分割成 n 等份
num = 20              # 螺線管匝數
mu = 1                # 真空中的磁導率
current = 5e3         # 電流量值
direct = True         # 電流方向, True 為姆指向右, Fasle 為姆指向左, 改變 segment.axis 計算方式
L = 100               # 畫面寬度
N = 10                # 將顯示的空間每邊切成 N 等份
Bmax = 50             # 顯示的磁場最大值
m = 1                 # 電子質量
q = -1                # 電子電量

"""
 2. 畫面設定
"""
# 產生動畫視窗
scene = canvas(title="Magnetic Mirror Effect", width=600, height=600, x=0, y=0, background=color.black)

# 產生空白串列 points, 在螺線管上等距離取點並填入 points 中
points1 = [sphere(pos=vector(L/2 - i*L/n, r*cos(2*pi/n*num*i), r*sin(2*pi/n*num*i)), radius=point_size, color=color.cyan) 
            for i in range(n)]
points2 = [sphere(pos=vector(200,0,0)+vector(L/2 - i*L/n, r*cos(2*pi/n*num*i), r*sin(2*pi/n*num*i)), radius=point_size, 
            color=color.cyan) for i in range(n)]

# 產生不同角度小球
def difdegree(v):
    evx=[]
    evy=[]
    energy=[]
    for i in range(0,360,5):
        evx.append(v*cos(2*pi*i/360))
        evy.append(v*sin(2*pi*i/360))
    electronic = [sphere(pos = vector(100,0,0),v=vector(evx[i],evy[i],0),a=vector(0,0,0),radius = 1.5,color = color.blue, 
                make_trail=True)for i in range(72)]
    for i in range(72):
        energy.append(0.5*m*((evx[i])**2+(evy[i])**2))
    return electronic,energy

# 改變視窗中心
scene.center=vector(100,0,0)

# 產生空白串列 segs, 從 points 依序一次讀取兩個點, 計算軸向量, 中點位置, 將螺線管切成很多小圓柱並填入 segs 中
segs1 = []
segs2 = []
for i in range(n-1):
    if(direct): dis = points1[i+1].pos - points1[i].pos
    else: dis = points1[i].pos - points1[i+1].pos
    mid = (points1[i+1].pos + points1[i].pos)/2
    segs1.append(cylinder(pos = mid, axis = dis, radius = seg_size, color = color.yellow))

for i in range(n-1):
    if(direct): dis = points2[i+1].pos - points2[i].pos
    else: dis = points2[i].pos - points2[i+1].pos
    mid = (points2[i+1].pos + points2[i].pos)/2
    segs2.append(cylinder(pos = mid, axis = dis, radius = seg_size, color = color.yellow))

# 自訂函式 magnetic, 計算某個位置的磁場
def magnetic(loc, segments):
    field = vector(0, 0, 0)
    for segment in segments:
        axis = loc - segment.pos
        field += mu*current/(4*pi)*mag(segment.axis)*cross(segment.axis, axis.norm())/mag(axis)**2
    return field  

# 計算羅倫茲力產生的加速度    
def lorentz(v,B):
    a=q*cross(v,B)/m
    return a

# 加入要畫箭頭的位置
locations = []
for i in range(N+1):
    for j in range(N+1):
        location = vector(100,0,0) + vector(L/N*i - L/2, L/N*j - L/2, 0)
        locations.append(location)

# 依序讀取串列 locations 的元素, 在對應的位置產生箭頭
fields = [arrow(pos=location, axis=vector(0, 0, 0), color=color.green) for location in locations]

# 更新箭頭的長度及方向, 若磁場量值 >= Bmax 則設定為 Bmax, 以避免箭頭蓋住其它東西
# 量值接近 Bmax 偏紅色, 量值接近 0 偏綠色
for field in fields:
    value = magnetic(field.pos, segs1)+ magnetic(field.pos, segs2)
    if(value.mag >= Bmax): value = value/value.mag * Bmax
    field.axis = value/10
    field.color = vector(value.mag/Bmax, 1 - value.mag/Bmax, 0)


"""
 3. 統計實測
"""     
analyze = 360     # 循環試驗次數   
dt = 0.0001        # 時間間隔     
correctline = 2   # 正確列數
incorrectline = 2 # 不正確列數
energyline = 2    # 動能列數
# 建立統計表
wb = openpyxl.Workbook()
wb.create_sheet("datasheet", 0)
ws = wb.active
ws['A1'].value = "Correct"
ws['B1'].value = "Incorrect"
ws['C1'].value = "initenergy"
ws['D1'].value = "afterenergy"
wb.save("vpython_data.xlsx")

# 算出各初始動能
electronic,energy = difdegree(30)
for i in range(72):
    ws[f'C{i+2}'].value = 0.5*m*((electronic[i].v.x)**2+(electronic[i].v.y)**2+(electronic[i].v.z)**2)

for i in range(72):
    if i>-1: 
        rate(100/dt)
        t=0
    # print(electronic[i].v.x,electronic[i].v.y,electronic[i].v.z)
        while t<5:
            electronic[i].a=lorentz(electronic[i].v,magnetic(electronic[i].pos,segs1)+magnetic(electronic[i].pos,segs2))
            electronic[i].v+=electronic[i].a*dt
            electronic[i].pos+=electronic[i].v*dt
            t+=dt

    # 截圖保存
        im = ImageGrab.grab(bbox=(0,50,1000,1000))  # x1 y1 x2 y2
    # 儲存檔案
        im.save(f'degree{i}.png')

        ws[f'D{energyline}'].value = 0.5*m*((electronic[i].v.x)**2+(electronic[i].v.y)**2+(electronic[i].v.z)**2)
        ws[f'E{energyline}'].value = f'{i} degree'
    # print(electronic[i].v.x,electronic[i].v.y,electronic[i].v.z)
        energyline+=1
    # 還在此範圍視為達到磁瓶效果
        if electronic[i].pos.x>50 and electronic[i].pos.x<150 and electronic[i].pos.y<75 and electronic[i].pos.y>-75:
            ws[f'A{correctline}'].value = f'{i} degree'
            wb.save("vpython_data.xlsx")
            correctline+=1
        else: 
            ws[f'B{incorrectline}'].value = f'{i} degree'
            wb.save("vpython_data.xlsx")
            incorrectline+=1
        electronic[i].clear_trail()
        electronic[i].opacity = 0
